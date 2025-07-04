# Taxonomy classification using MegaBlast and clean-up (python script):

# Merge qiime formatted FASTA sequence and taxonomy files
from Bio import SeqIO

# Paths to your input files
fasta_file = "COInr_qiime_animalia_all_trainseq.fasta"
taxonomy_file = "COInr_qiime_animalia_all_taxon.txt"
output_fasta_file = "COInr_qiime_animalia_all_trainseq_modified.fasta"

# Parse the taxonomy file into a dictionary
taxonomy_dict = {}
with open(taxonomy_file, 'r') as tax_file:
    for line in tax_file:
        parts = line.strip().split("\t")
        if len(parts) == 2:
            # Remove spaces from the taxonomic classification
            taxonomy_dict[parts[0]] = parts[1].replace(" ", "")

# Read the FASTA file and replace headers
with open(output_fasta_file, 'w') as output_fasta:
    for record in SeqIO.parse(fasta_file, "fasta"):
        # Replace the header with only the taxonomy if it exists in the taxonomy dictionary
        if record.id in taxonomy_dict:
            taxonomy = taxonomy_dict[record.id]
            record.id = taxonomy  # Set the taxonomic classification as the new ID
            record.description = ""  # Clear the description to avoid duplication
        # Write the updated record to the output file
        SeqIO.write(record, output_fasta, "fasta")

print(f"Modified FASTA file written to: '{output_fasta_file}'.")


# Generate mkCOInr blast database files
import subprocess

# Define the bash script as a string
bash_script = """
#!/bin/bash
makeblastdb -in COInr_qiime_animalia_all_trainseq_modified.fasta -input_type fasta -out mkCOInr_COI_Qiime_Animalia_derepl/mkCOInr_COI_Qiime_Animalia_derepl -dbtype nucl
"""
# Execute the bash script
try:
    result = subprocess.run(["bash", "-c", bash_script], capture_output=True, text=True, check=True)
    print("Bash Script Output:\n", result.stdout)
    print("Database files generated and stored in 'mkCOInr_COI_Qiime_Animalia_derepl' folder.")
except subprocess.CalledProcessError as e:
    print("Error occurred:\n", e.stderr)


# Run Megablast search against mkCOInr database
# Define the bash script as a string
bash_script = """
#!/bin/bash
blastn -db mkCOInr_COI_Qiime_Animalia_derepl/mkCOInr_COI_Qiime_Animalia_derepl -query dna-sequences.fasta -outfmt "10 qaccver saccver pident qcovhsp evalue" -perc_identity 98 -qcov_hsp_perc 90 -out mkCOInr_COI_Qiime_Animalia_derepl_all_sig_hits.csv
"""

# Execute the bash script
try:
    result = subprocess.run(["bash", "-c", bash_script], capture_output=True, text=True, check=True)
    print("Bash Script Output:\n", result.stdout)
    print("Megablast results saved in 'mkCOInr_COI_Qiime_Animalia_derepl_all_sig_hits.csv'.")
except subprocess.CalledProcessError as e:
    print("Error occurred:\n", e.stderr)


# Megablast results processing steps:

# Remove duplicate feature IDs and taxons
import pandas as pd
import numpy as np
import os

# Load the file directly (replace 'file_object' with the actual file object)
data = pd.read_csv('mkCOInr_COI_Qiime_Animalia_derepl_all_sig_hits.csv', header=None)

# Remove duplicates based on columns 0 and 1
cleaned_data = data.drop_duplicates(subset=[0, 1])

# Sort the cleaned data by column 0
sorted_data = cleaned_data.sort_values(by=0)

# Save the sorted and cleaned data to a new CSV file
sorted_data.to_csv("deduped_data.csv", index=False, header=False)

print("Sorted and deduped data has been saved to 'deduped_data.csv'.")


# Split taxons to multiple columns and clean
# Step 1: Read the CSV file (without header)
df = pd.read_csv('deduped_data.csv', header=None)

# Step 2: Split Column 2 (index 1) by semicolon and expand into separate columns
df_split = df[1].str.split(';', expand=True)

# Step 3: Clean up only columns 2 through 7 (index 1 through 6)
for col in range(0, 6):  # Loop through columns 2 to 7 (index 1 to 6)
    # Check for "__" and remove anything before it
    df_split[col] = df_split[col].str.split('__').str[1]
    # Remove anything after the first "_"
    df_split[col] = df_split[col].str.split('_').str[0]

# Step 4: Combine the original Column 1 (index 0) with the cleaned split columns and Columns 3, 4, and 5
df_combined = pd.concat([df[0], df_split, df[2], df[3], df[4]], axis=1)

# Step 5: Save the result to a new CSV file
df_combined.to_csv('deduped_data_splitted_cleaned.csv', index=False, header=False)

# Print a message indicating the completion of the task
print("The deduped data has been processed and saved to 'deduped_data_splitted_cleaned.csv'.")


# Remove underscore and clean species column
# Step 1: Read the CSV file (without header)
df = pd.read_csv('deduped_data_splitted_cleaned.csv', header=None)

# Step 6: Check if column 8 (index 7) exists before processing it
if 7 in df:
    df[7] = df[7].str.split('__').str[1]

# Step 7: Save the result to a new CSV file
df.to_csv('deduped_data_species_splitted_cleaned.csv', index=False, header=False)

# Print a message indicating the completion of the task
print("The species column of deduped data has been processed and saved to 'deduped_data_species_splitted_cleaned.csv'.")

# Function to remove anything after the second underscore
def remove_after_second_underscore(s):
    # Ensure s is a string
    if isinstance(s, str):
        # Find the first underscore
        first_underscore_index = s.find('_')
        
        # If the first underscore exists, find the second one starting after the first
        if first_underscore_index != -1:
            second_underscore_index = s.find('_', first_underscore_index + 1)
            if second_underscore_index != -1:
                # Return the substring up to the second underscore
                return s[:second_underscore_index]
    # If there are fewer than two underscores or s is not a string, return the original string
    return s

# Step 1: Read the CSV file (without header)
df = pd.read_csv('deduped_data_species_splitted_cleaned.csv', header=None)

# Step 2: Apply the function to column 8 (index 7) to remove anything after the second underscore
df[7] = df[7].apply(lambda x: remove_after_second_underscore(x))

# ADDITIONAL LINE: Replace underscores with spaces in column 8 (index 7)
df[7] = df[7].str.replace('_', ' ')

# Step 3: Save the result to a new CSV file
df.to_csv('deduped_data_all_splitted_cleaned.csv', index=False, header=False)

# Print a message indicating the completion of the task
print("All columns of deduped data has been processed and saved to 'deduped_data_all_splitted_cleaned.csv'.")


# Remove redundant taxons
from itertools import combinations

def is_subsequence(shorter, longer):
    """
    Check if 'shorter' is a subsequence of 'longer' (order-preserving but not necessarily contiguous).
    """
    it = iter(longer)
    return all(item in it for item in shorter if pd.notna(item))

def remove_redundant_rows_based_on_subsequence(data, columns):
    """
    Removes rows where the values in the specified columns are a subsequence of another row.
    """
    data_array = data.iloc[:, columns].to_numpy()
    to_keep = np.ones(len(data), dtype=bool)  # Boolean mask for rows to keep
    
    for i, j in combinations(range(len(data)), 2):
        row_i, row_j = data_array[i], data_array[j]

        if is_subsequence(row_i, row_j):  
            to_keep[i] = False  # row_i is redundant
        elif is_subsequence(row_j, row_i):
            to_keep[j] = False  # row_j is redundant

    return data[to_keep]

def remove_redundant_rows_based_on_column_0(data):
    """
    Removes redundant rows within groups defined by duplicates in column 0.
    """
    # Remove rows where columns 5,6,7 are blank
    data = data.dropna(subset=[5, 6, 7], how='all')  # Removes rows where all three columns are NaN

    cleaned_groups = []
    for _, group in data.groupby(0):  
        group_cleaned = remove_redundant_rows_based_on_subsequence(group, list(range(8)))
        if not group_cleaned.empty:
            cleaned_groups.append(group_cleaned)
    
    return pd.concat(cleaned_groups, ignore_index=True) if cleaned_groups else pd.DataFrame(columns=data.columns)

# Example usage
if __name__ == "__main__":
    data = pd.read_csv("deduped_data_all_splitted_cleaned.csv", header=None)

    cleaned_data = remove_redundant_rows_based_on_column_0(data)

    cleaned_data.to_csv("unique_data_all_splitted_cleaned.csv", index=False, header=False)
    print("Redundant rows removed. Unique data saved to 'unique_data_all_splitted_cleaned.csv'.")


# Remove taxons with percent identity difference of >0.5 from top hit
def filter_max_difference(csv_file):
    """
    Reads 'csv_file' (no header) into a DataFrame. Groups by values of column 0 (index 0),
    and for each group, retains rows whose column 8 (index 8) is within 0.5 of the max.
    Works on pandas versions that do NOT support 'include_groups'.
    """
    # 1. Load the CSV file without headers
    data = pd.read_csv(csv_file, header=None)

    # 2. Convert column 8 to numeric
    data[8] = pd.to_numeric(data[8], errors='coerce')

    # 3. Separate the grouping column (index 0)
    group_col = data[0]          # Series of group labels
    data_no_group = data.drop(columns=0)  # Remove column 0 from the DataFrame

    # 4. Define a function to filter within each group
    def filter_group(group_df, group_val):
        # group_df has columns [1, 2, 3, ... 8, ...] but no column 0
        max_value = group_df[8].max()
        # Keep rows where column 8 >= max_value - 0.5
        filtered = group_df[group_df[8] >= (max_value - 0.5)].copy()
        # If you need to re-add the group label as column 0, do so here:
        filtered.insert(0, 0, group_val)
        return filtered

    # 5. Group by the 'group_col' (the Series), not by a column in data_no_group
    filtered_data = (
        data_no_group
        .groupby(group_col, group_keys=False)
        .apply(lambda g: filter_group(g, g.name))
    )

    return filtered_data

# Example usage
if __name__ == "__main__":
    csv_file = "unique_data_all_splitted_cleaned.csv"
    filtered_data = filter_max_difference(csv_file)

    # Save or display the filtered data
    filtered_data.to_csv("filtered_data_all_splitted_cleaned.csv", index=False, header=False)
    print("Rows within 0.5 of the max in column 8 retained per group. Saved to 'filtered_data_all_splitted_cleaned.csv'.")


# Add header rows
# Specify the file paths
input_file = "filtered_data_all_splitted_cleaned.csv"  # Input CSV file without a header
output_file = "filtered_data_all_splitted_cleaned_with_header.csv"  # Output CSV file with the header

# Define the new header row
new_header = [
    "Feature ID",
    "mkCOInr kingdom name",
    "mkCOInr phylum name",
    "mkCOInr class name",
    "mkCOInr order name",
    "mkCOInr family name",
    "mkCOInr genus name",
    "mkCOInr species name",
    "Identity",
    "Coverage",
    "e-value"
]

# Read the input CSV file without a header
data = pd.read_csv(input_file, header=None)

# Assign the new header to the dataframe
data.columns = new_header

# Save the dataframe to a new CSV file with the header
data.to_csv(output_file, index=False)

print(f"The file has been updated with header row and saved as {output_file}.")


# Look-up against species list and perform occurence check
# Load the datasets
filtered_output_file = "filtered_data_all_splitted_cleaned_with_header.csv"  # Input file
state_species_file = "State_animalia_species_list.tsv"  # Reference file

# Read the input files
filtered_data = pd.read_csv(filtered_output_file)
state_species_list = pd.read_csv(state_species_file, sep="\t")

# Perform lookups and add occurrence columns
filtered_data['GBIF family occurrence'] = filtered_data['mkCOInr family name'].apply(
    lambda family_name: (
        '' if pd.isna(family_name) or family_name.strip() == '' else
        'PRESENT' if family_name in state_species_list['family'].values else
        'ABSENT'
    )
)
filtered_data['GBIF genus occurrence'] = filtered_data['mkCOInr genus name'].apply(
    lambda genus_name: (
        '' if pd.isna(genus_name) or genus_name.strip() == '' else
        'PRESENT' if genus_name in state_species_list['genus'].values else
        'ABSENT'
    )
)
filtered_data['GBIF species occurrence'] = filtered_data['mkCOInr species name'].apply(
    lambda species_name: (
        '' if pd.isna(species_name) or species_name.strip() == '' else
        'PRESENT' if species_name in state_species_list['species'].values else
        'ABSENT'
    )
)
# Rearrange the columns to place the new ones in the desired order
columns_to_rearrange = ['GBIF family occurrence', 'GBIF genus occurrence', 'GBIF species occurrence']
rearranged_columns = (
    filtered_data.columns.difference(columns_to_rearrange, sort=False).tolist() + columns_to_rearrange
)

filtered_data = filtered_data[rearranged_columns]

# Save the updated dataset to an Excel file
output_file = "filtered_output_with_gbif_occurrence.xlsx"
filtered_data.to_excel(output_file, index=False)

print(f"Updated dataset saved to {output_file}")


# Sort the duplicate feature IDs at top, unique feature IDs at bottom, and add empty 'Vetted ID' and 'Comments' columns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Input and output file paths
input_file = "filtered_output_with_gbif_occurrence.xlsx"
output_file = "processed_data_with_gbif_occurrence_highlighted_styled.xlsx"

# ---------------------------
# STEP 1: Load and Modify Data
# ---------------------------
# Load the Excel file into a DataFrame
df = pd.read_excel(input_file)

# Ensure at least 8 columns exist before adding new ones
if len(df.columns) < 8:
    raise ValueError("The dataset does not have at least 8 columns required for inserting new ones!")

# Convert 'Feature ID' to string type for consistency
df['Feature ID'] = df['Feature ID'].astype(str)

# Compute occurrence counts for Feature IDs
feature_counts = df['Feature ID'].value_counts()

# Mark duplicates (any Feature ID that appears more than once)
df['is_duplicate'] = df['Feature ID'].map(lambda x: feature_counts[x] > 1)

# **Sorting:**
# 1. Place all duplicate Feature IDs at the top
# 2. Within duplicates, group them by Feature ID
# 3. Place all unique Feature IDs at the bottom
df_sorted = df.sort_values(by=['is_duplicate', 'Feature ID'], ascending=[False, True]).drop(columns=['is_duplicate'])

# **Insert two new empty columns ("Vetted ID" and "Comments") after column 8**
df_columns = list(df_sorted.columns)
if len(df_columns) >= 8:
    col_position = 8  # Column index after which the new columns will be inserted
    df_sorted.insert(col_position, "Vetted ID", "")  # Add empty "Vetted ID" column
    df_sorted.insert(col_position + 1, "Comments", "")  # Add empty "Comments" column

# Save the modified DataFrame to a new Excel file
df_sorted.to_excel(output_file, index=False, sheet_name='Taxonomy', engine='openpyxl')


# Add conditional formatting colors to occurrence check columns and feature IDs
# -------------------------------
# STEP 2: Apply Excel Formatting
# -------------------------------
# Load workbook for formatting
wb = load_workbook(output_file)
ws = wb.active

# Define alternating color fills for duplicate groups (Light Gray & Light Salmon)
duplicate_fills = [
    PatternFill(start_color="FFE3D7", end_color="FFE3D7", fill_type="solid"),  # Pale Salmon
    PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),  # Light Salmon
]
# Define alternating color fills for unique IDs
unique_fills = [
    PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Light Blue
    PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),  # Lavender
]
present_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")     # Green for GBIF 'PRESENT'
absent_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")      # Red for GBIF 'ABSENT'
bold_font = Font(bold=True)
center_alignment = Alignment(wrap_text=True, horizontal="center")

# Apply bold and center alignment to the header row
for cell in ws[1]:
    cell.font = bold_font
    cell.alignment = center_alignment

# Get column index for "Feature ID" (Excel is 1-based)
feature_id_col = list(df_sorted.columns).index("Feature ID") + 1

# Apply alternating colors to distinguish different duplicate groups and unique IDs
prev_feature_id = None
duplicate_color_index = 0
unique_color_index = 0

for row_idx in range(2, ws.max_row + 1):  # Skip header row
    feature_id = ws.cell(row=row_idx, column=feature_id_col).value

    # Determine if the Feature ID is duplicate or unique
    if feature_counts[feature_id] > 1:  # Duplicates
        if feature_id != prev_feature_id:
            duplicate_color_index = (duplicate_color_index + 1) % len(duplicate_fills)  # Rotate colors
        ws.cell(row=row_idx, column=feature_id_col).fill = duplicate_fills[duplicate_color_index]
    else:  # Unique IDs
        unique_color_index = (unique_color_index + 1) % len(unique_fills)  # Rotate colors
        ws.cell(row=row_idx, column=feature_id_col).fill = unique_fills[unique_color_index]

    prev_feature_id = feature_id  # Track last seen Feature ID

# Identify GBIF occurrence columns
gbif_columns = ["GBIF family occurrence", "GBIF genus occurrence", "GBIF species occurrence"]
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
gbif_indices = [header_row.index(col) + 1 for col in gbif_columns if col in header_row]

# Apply formatting to GBIF occurrence columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for idx in gbif_indices:
        cell = row[idx - 1]
        if cell.value and isinstance(cell.value, str):
            value = cell.value.strip().upper()
            if value == "PRESENT":
                cell.fill = present_fill
            elif value == "ABSENT":
                cell.fill = absent_fill

# Save the formatted Excel file
wb.save(output_file)

print(f"✅ Sorted and formatted file saved as: {output_file}")

os.remove('deduped_data.csv')
os.remove('deduped_data_splitted_cleaned.csv')
os.remove('deduped_data_species_splitted_cleaned.csv')
os.remove('deduped_data_all_splitted_cleaned.csv')
os.remove('unique_data_all_splitted_cleaned.csv')
os.remove('filtered_data_all_splitted_cleaned.csv')
os.remove('filtered_data_all_splitted_cleaned_with_header.csv')
os.remove('filtered_output_with_gbif_occurrence.xlsx')


# Consolidate metadata post vetting (python script):

# Remove duplicate Feature IDs and Vetted ID, fill Vetted ID, rename 'Vetted ID' as 'Final taxon name', and remove occurrence check columns
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

### STEP 1: Process the original Excel file and create 'Vetted_Hawaii_pollinator_2024_taxonomy.csv'

# Load the Excel file
excel_file = "Hawaii_pollinator_2024_taxonomy_to_vet_MDJ_test.xlsx"
df = pd.read_excel(excel_file)

# Drop duplicates based on 'Feature ID' and 'Vetted ID'
df_unique = df.drop_duplicates(subset=['Feature ID', 'Vetted ID'])

# Drop specified columns
columns_to_drop = ['Comments', 'GBIF family occurrence', 'GBIF genus occurrence', 'GBIF species occurrence']
df_unique = df_unique.drop(columns=columns_to_drop, errors='ignore')

# Fill missing 'Vetted ID' with hierarchical replacement
df_unique['Vetted ID'] = df_unique['Vetted ID'].fillna(df_unique['mkCOInr species name'])
df_unique['Vetted ID'] = df_unique['Vetted ID'].fillna(df_unique['mkCOInr genus name'])
df_unique['Vetted ID'] = df_unique['Vetted ID'].fillna(df_unique['mkCOInr family name'])

# Rename 'Vetted ID' to 'Final taxon name'
df_unique = df_unique.rename(columns={'Vetted ID': 'Final taxon name'})

# Save the cleaned data as a CSV file with the new name
csv_file = "Vetted_Hawaii_pollinator_2024_taxonomy.csv"
df_unique.to_csv(csv_file, index=False)
print(f"Processed file saved as: {csv_file}")


# Normalize feature table, add FASTA sequence, and merge with vetted taxonomy table
### STEP 2: Merge with FASTA and TSV data, and create the final Excel file

# Define file paths
fasta_file = "dna-sequences.fasta"  # FASTA file
tsv_file = "feature-table.tsv"  # TSV file

# Load the CSV file
df_csv = pd.read_csv(csv_file)

# Read the FASTA file into a dictionary
fasta_dict = {}
with open(fasta_file, "r") as file:
    seq_id = None
    sequence = ""
    for line in file:
        line = line.strip()
        if line.startswith(">"):  # Header line
            if seq_id:
                fasta_dict[seq_id] = sequence  # Store previous sequence
            seq_id = line[1:]  # Remove '>' from header
            sequence = ""
        else:
            sequence += line  # Append sequence data
    if seq_id:
        fasta_dict[seq_id] = sequence  # Store last sequence

# Create a new column in the CSV file for matched sequences
df_csv["FASTA Sequence"] = df_csv["Feature ID"].map(fasta_dict)

# Load the TSV file and remove the first row
df_tsv = pd.read_csv(tsv_file, sep="\t", skiprows=1)

# Identify columns that start with "Negative"
negative_cols = [col for col in df_tsv.columns if col.startswith("Negative")]
other_cols = [col for col in df_tsv.columns if col not in negative_cols]

# Reorder columns by moving negative columns to the right end
df_tsv = df_tsv[other_cols + negative_cols]

# Calculate the max of negative columns for each row
if negative_cols:
    df_tsv["Max_Negative"] = df_tsv[negative_cols].max(axis=1)
    
    # Subtract max negative value from all other columns
    for col in other_cols:
        if col != "#OTU ID":  # Ensure ID column remains unchanged
            df_tsv[col] = df_tsv[col] - df_tsv["Max_Negative"]
            df_tsv[col] = df_tsv[col].clip(lower=0)  # Replace negative values with zero
    
    # Drop the temporary max column
    df_tsv.drop(columns=["Max_Negative"], inplace=True)

# Merge based on '#OTU ID' in TSV and 'Feature ID' in CSV
df_merged = df_csv.merge(df_tsv, left_on='Feature ID', right_on='#OTU ID', how='left')

# Drop the '#OTU ID' column
df_merged.drop(columns=["#OTU ID"], inplace=True)

# Save the final merged data as an Excel file
excel_path = "Hawaii_invertebrate_2024_final_metadata.xlsx"
df_merged.to_excel(excel_path, index=False)

# Load the workbook using openpyxl to apply formatting
wb = load_workbook(excel_path)
ws = wb.active

# Rename the worksheet tab to 'Metadata'
ws.title = "Metadata"

# Set the header row font and wrap text
header_font = Font(name='Aptos Narrow', size=12, bold=True)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
 
# Apply font and alignment to header cells (header row is the first row)
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = header_alignment

# Save the formatted Excel file
wb.save(excel_path)

print("Processing complete. Final Excel file saved as 'Hawaii_invertebrate_2024_final_metadata.xlsx'.")